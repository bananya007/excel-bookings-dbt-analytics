"""Generate a reproducible legacy workbook for the spreadsheet-to-dbt project."""
import argparse
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)
REGIONS = ["East", "West", "Central", "South"]
REGION_VARIANTS = {"East": ["East", "East ", "east", "EAST"], "West": ["West", "west ", "West"], "Central": ["Central", "central", "Central "], "South": ["South", "SOUTH", "south"]}
PRODUCTS = ["CorePlatform", "AnalyticsAddon", "SecuritySuite", "DataConnect"]
DEAL_TYPES = ["New License", "Renewal", "Support", "Services"]
CURRENCIES = ["USD", "EUR", "GBP", "INR", "CAD", "BRL"]
RATE_VALUES = {"USD": 1.0, "EUR": 1.09, "GBP": 1.28, "INR": 0.012, "CAD": 0.74}
REPS = [("John Smith", "East", "REP001"), ("Maria Garcia", "West", "REP002"), ("Wei Chen", "Central", "REP003"), ("Priya Sharma", "South", "REP004"), ("Tom Becker", "East", "REP005"), ("Sara Cohen", "West", "REP006"), ("Luis Ortega", "Central", "REP007"), ("Anna Kowalski", "South", "REP008")]
ALIASES = {"John Smith": "Jon Smith", "Anna Kowalski": "Ana Kowalski"}
N_ROWS = 400
START_DATE = date(2024, 1, 1)

def build_workbook():
    wb = Workbook()
    raw = wb.active
    raw.title = "Raw_Bookings"
    headers = ["BookingID", "Booking Date", "Customer", "Region", "Product", "Deal Type", "Sales Rep", "Amount (Local)", "Currency"]
    raw.merge_cells("A1:I1")
    raw["A1"] = "RAW EXPORT - DO NOT EDIT (from CRM, monthly)"
    raw["A1"].font = Font(bold=True, italic=True, color="808080")
    for col, header in enumerate(headers, 1):
        raw.cell(2, col, header).font = Font(bold=True)
    rows = []
    for i in range(N_ROWS):
        d = START_DATE + timedelta(days=random.randint(0, 545))
        region = random.choice(REGIONS)
        rep = random.choice([n for n, r, _ in REPS if r == region])
        rows.append([f"BK{10000+i}", d, f"Customer {random.randint(1,120):03d}", random.choice(REGION_VARIANTS[region]), random.choice(PRODUCTS), random.choice(DEAL_TYPES), rep, round(random.uniform(3000,180000), 2), random.choices(CURRENCIES, weights=[40,20,12,12,10,6])[0]])
    for i in (25, 90, 210):
        rows.append(list(rows[i]))
    random.shuffle(rows)
    for excel_row, row in enumerate(rows, 3):
        for col, value in enumerate(row, 1):
            cell = raw.cell(excel_row, col, value)
            if col == 2:
                mode = random.random()
                if mode < 0.5:
                    cell.number_format = "mm/dd/yyyy"
                elif mode < 0.8:
                    cell.value = value.strftime("%m/%d/%Y")
                else:
                    cell.value = value.isoformat()
    raw_last = 2 + len(rows)

    rates = wb.create_sheet("Rates")
    for col, header in enumerate(["Currency", "Rate to USD", "Effective Date", "Source"], 1):
        rates.cell(1, col, header).font = Font(bold=True)
    for row, (currency, rate) in enumerate(RATE_VALUES.items(), 2):
        rates.cell(row, 1, currency); rates.cell(row, 2, rate)
        rates.cell(row, 3, date(2024, 1, 1)).number_format = "yyyy-mm-dd"
        rates.cell(row, 4, "Finance rate card")
    rates["F2"] = "Priya updates rates on the 1st. DO NOT TOUCH!!"
    rates["F4"] = "TODO: add BRL when Brazil deals close (ask Finance)"

    targets = wb.create_sheet("Rep_Targets")
    for col, header in enumerate(["Rep ID", "Sales Rep", "Target Year", "Annual Quota (USD)"], 1):
        targets.cell(1, col, header).font = Font(bold=True)
    for row, (name, _, rep_id) in enumerate(REPS, 2):
        targets.cell(row, 1, rep_id); targets.cell(row, 2, ALIASES.get(name, name))
        targets.cell(row, 3, 2024); targets.cell(row, 4, random.choice([900000, 1100000, 1250000, 1400000]))

    master = wb.create_sheet("Rep_Master")
    for col, header in enumerate(["Alias Name", "Canonical Name", "Rep ID", "Region"], 1):
        master.cell(1, col, header).font = Font(bold=True)
    row = 2
    for name, region, rep_id in REPS:
        for alias in [name, ALIASES.get(name)]:
            if alias:
                master.cell(row, 1, alias); master.cell(row, 2, name)
                master.cell(row, 3, rep_id); master.cell(row, 4, region); row += 1

    calc = wb.create_sheet("Calc")
    calc_headers = ["BookingID", "Clean Region", "Amount USD", "Deal Category", "Rep ID", "Rep Quota", "Attainment Status", "Adjusted USD", "Adjustment ID", "Adjustment Source"]
    for col, header in enumerate(calc_headers, 1):
        calc.cell(1, col, header).font = Font(bold=True)
        calc.cell(1, col).fill = PatternFill("solid", start_color="FCE4D6")
    calc_last = raw_last - 1
    adjustment_rows = set(random.sample(range(2, calc_last + 1), 5))
    for raw_row in range(3, raw_last + 1):
        calc_row = raw_row - 1
        calc.cell(calc_row, 1, f"=Raw_Bookings!A{raw_row}")
        calc.cell(calc_row, 2, f'=IF(TRIM(LOWER(Raw_Bookings!D{raw_row}))="east","East",IF(TRIM(LOWER(Raw_Bookings!D{raw_row}))="west","West",IF(TRIM(LOWER(Raw_Bookings!D{raw_row}))="central","Central",IF(TRIM(LOWER(Raw_Bookings!D{raw_row}))="south","South","UNKNOWN"))))')
        calc.cell(calc_row, 3, f"=IFERROR(Raw_Bookings!H{raw_row}*VLOOKUP(Raw_Bookings!I{raw_row},Rates!$A$2:$B$6,2,FALSE),Raw_Bookings!H{raw_row})")
        calc.cell(calc_row, 4, f'=IF(Raw_Bookings!F{raw_row}="New License","NEW",IF(OR(Raw_Bookings!F{raw_row}="Renewal",Raw_Bookings!F{raw_row}="Support"),"RECURRING","OTHER"))')
        calc.cell(calc_row, 5, f'=IFERROR(VLOOKUP(Raw_Bookings!G{raw_row},Rep_Master!$A$2:$D$18,3,FALSE),"")')
        calc.cell(calc_row, 6, f"=IFERROR(VLOOKUP(E{calc_row},Rep_Targets!$A$2:$D$9,4,FALSE),0)")
        calc.cell(calc_row, 7, f'=IF(F{calc_row}=0,"NO QUOTA FOUND",IF(C{calc_row}>=F{calc_row}*0.25,"On Track","Behind"))')
        if calc_row in adjustment_rows:
            calc.cell(calc_row, 8, round(random.uniform(5000, 90000), 2))
            calc.cell(calc_row, 9, f"ADJ{calc_row-1:03d}")
            calc.cell(calc_row, 10, "mgr adj - see email 3/12")
        else:
            calc.cell(calc_row, 8, f'=C{calc_row}*IF(B{calc_row}="West",0.95,1)')

    summary = wb.create_sheet("Regional_Summary")
    summary.merge_cells("A1:F1")
    summary["A1"] = "REGIONAL BOOKINGS SUMMARY - FINAL - v7"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A1"].alignment = Alignment(horizontal="center")
    for col, header in enumerate(["Region", "NEW", "RECURRING", "OTHER", "Total", "% of Grand Total"], 1):
        summary.cell(3, col, header).font = Font(bold=True)
    for row, region in enumerate(REGIONS, 4):
        summary.cell(row, 1, region)
        for col, category in enumerate(["NEW", "RECURRING", "OTHER"], 2):
            letter = get_column_letter(col)
            formula = "=SUMIFS(Calc!$H$2:$H{end},Calc!$B$2:$B{end},$A{row},Calc!$D$2:$D{end},{letter}$3)".format(end=calc_last, row=row, letter=letter)
            summary.cell(row, col, formula)
        summary.cell(row, 5, f"=SUM(B{row}:D{row})")
        summary.cell(row, 6, f"=IF($E$9=0,0,E{row}/$E$9)")
        summary.cell(row, 6).number_format = "0.0%"
    summary["A9"] = "GRAND TOTAL"; summary["E9"] = "=SUM(E4:E7)+5000"
    summary["E9"].font = Font(bold=True); summary["E9"].fill = PatternFill("solid", start_color="FFFF00")
    summary["G9"] = "true-up per Finance (Jan) - DO NOT REMOVE"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    return wb, len(rows), len(adjustment_rows)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    workbook, raw_count, adjustment_count = build_workbook()
    workbook.save(args.output)
    print(f"written: {raw_count} raw rows (incl. 3 duplicate IDs), {adjustment_count} hardcoded adjustments, 8 rep IDs")

if __name__ == "__main__":
    main()

